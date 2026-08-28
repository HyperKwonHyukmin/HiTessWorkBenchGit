"""이중관 PSA 어댑터(`hitess_adapter`) 테스트.

이 파일의 존재 이유는 두 가지다.

1. shim 4종이 실제로 의도한 대로 동작하는지 (엔진 무수정의 근거).
2. **드리프트 조기 감지** — 연구원이 새 엔진을 내려줬을 때 `pytest tests/test_doublepipe_adapter.py`
   만 돌리면 Tier-2 패치 앵커가 아직 유효한지 즉시 알 수 있다. 빨간 줄이 나면 patches.py 를
   갱신해야 한다는 뜻이다.
"""
import io
import locale
import os
import subprocess
import sys

import pytest

_BACKEND_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_ADAPTER_ROOT = os.path.join(_BACKEND_DIR, "InHouseAdapters", "doublepipe_psa")
if _ADAPTER_ROOT not in sys.path:
    sys.path.insert(0, _ADAPTER_ROOT)

from hitess_adapter import cli, engine, patches  # noqa: E402
from hitess_adapter.shims import abaqus_subprocess, openpyxl_drm, openpyxl_merged  # noqa: E402

openpyxl = pytest.importorskip("openpyxl")


# ---------------------------------------------------------------------------
# 공통 픽스처 — shim 은 전역 상태를 바꾸므로 테스트 후 반드시 원복한다.
# ---------------------------------------------------------------------------
@pytest.fixture
def merged_shim():
    from openpyxl.cell.cell import MergedCell

    original = MergedCell.__dict__.get("value")
    had_flag = getattr(MergedCell, "_hitess_merged_shim", False)
    openpyxl_merged._installed = False
    openpyxl_merged.install()
    yield
    MergedCell.value = original
    if not had_flag:
        try:
            del MergedCell._hitess_merged_shim
        except AttributeError:
            pass
    openpyxl_merged._installed = False


@pytest.fixture
def drm_shim():
    import openpyxl.reader.excel as reader

    original_pkg = openpyxl.load_workbook
    original_reader = reader.load_workbook
    openpyxl_drm._installed = False
    openpyxl_drm.install()
    yield
    openpyxl.load_workbook = original_pkg
    reader.load_workbook = original_reader
    openpyxl_drm._installed = False
    openpyxl_drm._orig_load_workbook = None


@pytest.fixture
def abaqus_shim():
    original = subprocess.Popen.__init__
    abaqus_subprocess._installed = False
    abaqus_subprocess.install()
    yield
    subprocess.Popen.__init__ = original
    abaqus_subprocess._installed = False


def _xlsx_bytes(value="hello"):
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = value
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# shim: MergedCell
# ---------------------------------------------------------------------------
def test_merged_cell_write_raises_without_shim():
    """shim 이 없으면 병합 비앵커 셀 쓰기는 죽는다 — 이것이 개조가 필요했던 이유."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.merge_cells("B13:H13")
    with pytest.raises(AttributeError):
        sheet["C13"].value = "x"


def test_merged_cell_write_is_ignored_with_shim(merged_shim):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.merge_cells("B13:H13")
    sheet["C13"].value = "x"          # 크래시 없이 무시되어야 한다
    assert sheet["C13"].value is None
    sheet["B13"].value = "anchor"     # 앵커 셀은 정상 기록
    assert sheet["B13"].value == "anchor"


def test_merged_cell_shim_is_idempotent(merged_shim):
    openpyxl_merged._installed = False
    openpyxl_merged.install()  # 두 번 설치해도 안전해야 한다
    workbook = openpyxl.Workbook()
    workbook.active.merge_cells("A1:C1")
    workbook.active["B1"].value = "x"
    assert workbook.active["B1"].value is None


# ---------------------------------------------------------------------------
# shim: openpyxl DRM 폴백
# ---------------------------------------------------------------------------
def test_drm_shim_loads_normal_xlsx(drm_shim, tmp_path):
    path = tmp_path / "plain.xlsx"
    path.write_bytes(_xlsx_bytes("ok"))
    workbook = openpyxl.load_workbook(str(path))
    assert workbook.active["A1"].value == "ok"


def test_drm_shim_falls_back_to_bundled_template(drm_shim, tmp_path, monkeypatch):
    """DRM 암호화된 서식 템플릿 → 번들 report_template.bin 으로 폴백."""
    monkeypatch.chdir(tmp_path)
    # HHIDRMC 헤더 + 쓰레기 = 회사 DRM 이 at-rest 암호화한 xlsx 의 모양
    (tmp_path / "Report for PSA.xlsx").write_bytes(b"HHIDRMC" + b"\x00" * 128)
    (tmp_path / "report_template.bin").write_bytes(_xlsx_bytes("template"))

    workbook = openpyxl.load_workbook("Report for PSA.xlsx")
    assert workbook.active["A1"].value == "template"


def test_drm_shim_reports_when_no_bundle(drm_shim, tmp_path, monkeypatch, capsys):
    """번들 사본조차 없으면 경고를 남기고 원래 예외가 올라가야 한다(조용한 성공 금지)."""
    monkeypatch.chdir(tmp_path)
    (tmp_path / "Report for PSA.xlsx").write_bytes(b"HHIDRMC" + b"\x00" * 128)
    with pytest.raises(Exception):
        openpyxl.load_workbook("Report for PSA.xlsx")
    assert "번들 사본" in capsys.readouterr().out


def test_drm_shim_does_not_swallow_other_broken_files(drm_shim, tmp_path):
    """서식 템플릿이 아닌 손상 파일은 폴백 없이 원래 예외로 올라가야 한다."""
    path = tmp_path / "some_other.xlsx"
    path.write_bytes(b"not a zip")
    with pytest.raises(Exception):
        openpyxl.load_workbook(str(path))


# ---------------------------------------------------------------------------
# shim: Abaqus subprocess
# ---------------------------------------------------------------------------
def test_ask_delete_is_injected_before_int():
    rewritten = abaqus_subprocess.rewrite_command("abaqus job=L17_SUS cpus=2 int")
    assert rewritten == "abaqus job=L17_SUS cpus=2 ask_delete=OFF int"


def test_ask_delete_is_not_duplicated():
    command = "abaqus job=L1 cpus=2 ask_delete=OFF int"
    assert abaqus_subprocess.rewrite_command(command) == command


def test_ask_delete_appends_when_no_int_tail():
    assert abaqus_subprocess.rewrite_command("abaqus job=L1 cpus=2") == "abaqus job=L1 cpus=2 ask_delete=OFF"


def test_non_abaqus_command_is_untouched():
    for command in ("python foo.py", ["abaqus", "job=L1"], "nastran x.bdf int"):
        assert abaqus_subprocess.rewrite_command(command) == command


def test_encoding_is_corrected_only_for_abaqus():
    fixed = abaqus_subprocess.fix_encoding("abaqus job=L1 int", {"encoding": "utf-8", "shell": True})
    assert fixed["encoding"] == locale.getpreferredencoding(False)

    untouched = abaqus_subprocess.fix_encoding("python x.py", {"encoding": "utf-8"})
    assert untouched["encoding"] == "utf-8"


def test_popen_shim_rewrites_actual_call(abaqus_shim, monkeypatch):
    """Popen 을 실제로 감쌌는지 — 명령이 재작성된 채 원래 __init__ 로 넘어가야 한다."""
    seen = {}

    def _fake_init(self, args, *rest, **kwargs):
        seen["args"] = args
        seen["encoding"] = kwargs.get("encoding")

    monkeypatch.setattr(abaqus_subprocess, "_orig_popen_init", _fake_init)
    subprocess.Popen("abaqus job=L2 cpus=2 int", shell=True, encoding="utf-8")

    assert seen["args"] == "abaqus job=L2 cpus=2 ask_delete=OFF int"
    assert seen["encoding"] == locale.getpreferredencoding(False)


# ---------------------------------------------------------------------------
# CLI — 기존 exe 와 동일한 인자 규약을 유지해야 한다
# ---------------------------------------------------------------------------
def test_parse_cases_always_includes_l17():
    assert cli.parse_cases(["L20"]) == {"L20", "L17"}
    assert cli.parse_cases(["L20,L21", "22"]) == {"L20", "L21", "L22", "L17"}
    assert cli.parse_cases(["l18"]) == {"L18", "L17"}


def test_parse_cases_rejects_bad_tokens():
    with pytest.raises(SystemExit):
        cli.parse_cases(["LX"])
    with pytest.raises(SystemExit):
        cli.parse_cases(["L30"])
    with pytest.raises(SystemExit):
        cli.parse_cases([""])


def test_select_indices_preserves_ascending_order():
    names = [f"L{n}_OPE_x" for n in range(1, 30)]
    assert cli.select_indices(names, {"L17", "L20"}) == [16, 19]


def test_parser_accepts_backend_command_shape():
    """백엔드가 만드는 command = [exe, csv] (+ --load-cases 'L18,L20') 형태."""
    args = cli.build_parser().parse_args(["C:/jobs/x.csv", "--load-cases", "L18,L20"])
    assert args.csv == "C:/jobs/x.csv"
    assert cli.parse_cases(args.load_cases) == {"L18", "L20", "L17"}

    default = cli.build_parser().parse_args(["C:/jobs/x.csv"])
    assert default.load_cases is None   # 미지정 = 전체 29개


# ---------------------------------------------------------------------------
# Tier-2 패치 — ★ 새 엔진이 왔을 때 여기가 먼저 빨개진다
# ---------------------------------------------------------------------------
def test_patch_apply_and_idempotency():
    patch = patches.Patch(file="x.py", why="테스트", anchor="AAA", replace="BBB", expect=1)
    assert patches.apply_to_text("x AAA y", patch) == ("x BBB y", "applied")
    assert patches.apply_to_text("x BBB y", patch) == ("x BBB y", "already")


def test_patch_raises_when_anchor_missing():
    patch = patches.Patch(file="x.py", why="테스트", anchor="AAA", replace="BBB", expect=1)
    with pytest.raises(patches.PatchError):
        patches.apply_to_text("nothing here", patch)


def test_patch_raises_on_unexpected_match_count():
    patch = patches.Patch(file="x.py", why="테스트", anchor="AAA", replace="BBB", expect=1)
    with pytest.raises(patches.PatchError):
        patches.apply_to_text("AAA AAA", patch)


@pytest.mark.parametrize("patch", patches.PATCHES, ids=lambda p: p.file)
def test_patch_still_resolvable_against_current_engine(patch):
    """현재 엔진 소스에 대해 패치가 여전히 '적용 가능' 하거나 '이미 반영됨' 이어야 한다.

    실패하면 연구원이 해당 코드를 바꾼 것이다 → patches.py 를 갱신하고 exe 를 재빌드할 것.
    """
    engine_dir = engine.default_engine_dir()
    target = os.path.join(engine_dir, patch.file)
    if not os.path.isfile(target):
        pytest.skip(f"엔진 소스 없음(빌드 환경 아님): {target}")
    with open(target, "r", encoding="utf-8") as handle:
        text = handle.read()
    _, status = patches.apply_to_text(text, patch)   # PatchError 면 테스트 실패
    assert status in ("applied", "already")


# ---------------------------------------------------------------------------
# 엔진 계약 검증
# ---------------------------------------------------------------------------
def test_engine_load_reports_missing_symbols(monkeypatch):
    """엔진에서 심볼이 사라지면 이름을 찍어 즉시 실패해야 한다."""
    import types

    monkeypatch.setattr(engine, "ensure_engine_on_path", lambda: None)
    stub_creator = types.ModuleType("AbaqusModelCreator")   # AbaqusModelCreator 심볼 없음
    stub_head = types.ModuleType("Head_for_FuelLine_ASME_B313_v2018")
    monkeypatch.setitem(sys.modules, "AbaqusModelCreator", stub_creator)
    monkeypatch.setitem(sys.modules, "Head_for_FuelLine_ASME_B313_v2018", stub_head)

    with pytest.raises(engine.EngineContractError) as excinfo:
        engine.load()
    message = str(excinfo.value)
    assert "AbaqusModelCreator.AbaqusModelCreator" in message
    assert "Head_for_FuelLine_ASME_B313_v2018.make_report" in message


def test_verify_head_instance_reports_missing_attrs():
    class Incomplete:
        LC_name = []

    with pytest.raises(engine.EngineContractError) as excinfo:
        engine.verify_head_instance(Incomplete())
    assert "datname" in str(excinfo.value)


def test_verify_head_instance_accepts_complete():
    class Complete:
        LC_name = []
        LC_name_iter01 = []
        LC_name_iter02 = []
        filename = []
        datname = []

        def Abaqus_Run(self, names, message=""):
            return None

    instance = Complete()
    assert engine.verify_head_instance(instance) is instance


# ---------------------------------------------------------------------------
# 백엔드 경로 해석 — 어댑터 폴더 우선, 구 배치 폴백
# ---------------------------------------------------------------------------
def test_backend_prefers_adapter_dir(monkeypatch, tmp_path):
    from app.services import doublepipe_psa_service as service

    adapter_dir = tmp_path / "HiTessAdapter"
    legacy_dir = tmp_path / "engine"
    adapter_dir.mkdir()
    legacy_dir.mkdir()
    (adapter_dir / service._PSA_EXE_NAME).write_bytes(b"new")
    (legacy_dir / service._PSA_EXE_NAME).write_bytes(b"old")
    monkeypatch.setattr(service, "_PSA_DIR", str(adapter_dir))
    monkeypatch.setattr(service, "_PSA_LEGACY_DIR", str(legacy_dir))

    assert service._resolve_psa_exe() == str(adapter_dir / service._PSA_EXE_NAME)


def test_backend_falls_back_to_legacy_dir(monkeypatch, tmp_path):
    from app.services import doublepipe_psa_service as service

    adapter_dir = tmp_path / "HiTessAdapter"
    legacy_dir = tmp_path / "engine"
    adapter_dir.mkdir()
    legacy_dir.mkdir()
    (legacy_dir / service._PSA_EXE_NAME).write_bytes(b"old")
    monkeypatch.setattr(service, "_PSA_DIR", str(adapter_dir))
    monkeypatch.setattr(service, "_PSA_LEGACY_DIR", str(legacy_dir))

    assert service._resolve_psa_exe() == str(legacy_dir / service._PSA_EXE_NAME)


def test_backend_returns_none_when_missing(monkeypatch, tmp_path):
    from app.services import doublepipe_psa_service as service

    monkeypatch.setattr(service, "_PSA_DIR", str(tmp_path / "nope"))
    monkeypatch.setattr(service, "_PSA_LEGACY_DIR", str(tmp_path / "also-nope"))
    assert service._resolve_psa_exe() is None
