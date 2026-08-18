"""GenericRenderer — ReportDoc 을 XLSX bytes 로."""
import io
import zipfile

import pytest

import openpyxl

from app.services.report.models import (
    ReportDoc,
    ReportField,
    ReportMeta,
    ReportSection,
    ReportTable,
)
from app.services.report.renderers.generic_xlsx import _text_width, render_generic_xlsx


def _doc(**overrides) -> ReportDoc:
    base = dict(
        meta=ReportMeta(
            program_id="column-buckling",
            display_name="Column Buckling Load Calculator",
            analysis_id=7,
            project_name="ColumnBuckling_20260818",
            employee_id="EMP001",
            created_at=None,
            status="Success",
        ),
        verdict="합격",
        sections=(
            ReportSection(
                key="overview",
                title="개요",
                fields=(ReportField(label="해석 App", value="Column Buckling Load Calculator"),),
            ),
            ReportSection(
                key="result",
                title="해석 결과",
                fields=(ReportField(label="최대 허용 하중", value=12.5, unit="ton"),),
                tables=(
                    ReportTable(
                        title="후보",
                        columns=("호칭", "중량"),
                        rows=((1, 2.0), (3, 4.0)),
                    ),
                ),
            ),
        ),
    )
    base.update(overrides)
    return ReportDoc(**base)


def _load(data: bytes):
    return openpyxl.load_workbook(io.BytesIO(data))


def test_returns_a_valid_xlsx_zip():
    data = render_generic_xlsx(_doc())
    assert zipfile.is_zipfile(io.BytesIO(data))


def test_creates_one_sheet_per_section_plus_cover():
    wb = _load(render_generic_xlsx(_doc()))
    assert wb.sheetnames[0] == "표지"
    assert "개요" in wb.sheetnames
    assert "해석 결과" in wb.sheetnames


def test_cover_shows_verdict_as_text():
    wb = _load(render_generic_xlsx(_doc()))
    cover = wb["표지"]
    values = [cell.value for row in cover.iter_rows() for cell in row]
    assert "합격" in values


def test_field_unit_lands_in_its_own_column():
    wb = _load(render_generic_xlsx(_doc()))
    ws = wb["해석 결과"]
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    assert ["최대 허용 하중", 12.5, "ton"] in [row[:3] for row in rows]


def test_table_headers_and_rows_are_written():
    wb = _load(render_generic_xlsx(_doc()))
    ws = wb["해석 결과"]
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    flat = [row[:2] for row in rows]
    assert ["호칭", "중량"] in flat
    assert [1, 2.0] in flat


def test_notices_are_written_on_the_cover():
    doc = _doc(notices=("표준 양식 미적용",))
    wb = _load(render_generic_xlsx(doc))
    values = [cell.value for row in wb["표지"].iter_rows() for cell in row]
    assert "표준 양식 미적용" in values


def test_duplicate_section_titles_get_unique_sheet_names():
    doc = _doc(
        sections=(
            ReportSection(key="overview", title="개요"),
            ReportSection(key="custom", title="개요"),
        )
    )
    wb = _load(render_generic_xlsx(doc))
    assert len(set(wb.sheetnames)) == len(wb.sheetnames)


def test_long_section_title_is_truncated_to_excel_sheet_limit():
    doc = _doc(sections=(ReportSection(key="overview", title="가" * 40),))
    wb = _load(render_generic_xlsx(doc))
    assert all(len(name) <= 31 for name in wb.sheetnames)


# ── 판정을 읽히게 만들기 ────────────────────────────────────────────────
# 결재자가 긴 표를 훑을 때 실패 행이 눈에 걸려야 한다. 색만으로 알리지는 않는다
# (글자는 그대로 '불합격'), 색을 아예 안 쓰지도 않는다.

def _cells(ws):
    return [cell for row in ws.iter_rows() for cell in row]


def _table_doc(rows, columns=("부재", "판정")):
    return _doc(
        sections=(
            ReportSection(
                key="result",
                title="해석 결과",
                tables=(ReportTable(title="부재", columns=columns, rows=rows),),
            ),
        )
    )


def test_failing_table_row_is_filled_and_passing_row_is_not():
    ws = _load(render_generic_xlsx(_table_doc((("A", "합격"), ("B", "불합격")))))["해석 결과"]

    fail_cell = next(c for c in _cells(ws) if c.value == "불합격")
    pass_cell = next(c for c in _cells(ws) if c.value == "합격")

    assert fail_cell.fill.patternType == "solid"
    assert fail_cell.font.color.rgb.endswith("CC0000")
    assert pass_cell.fill.patternType is None


def test_verdict_cell_on_cover_is_coloured_as_well_as_written():
    ws = _load(render_generic_xlsx(_doc(verdict="불합격")))["표지"]
    cell = next(c for c in _cells(ws) if c.value == "불합격")
    assert cell.font.color.rgb.endswith("CC0000")


def test_wide_table_columns_are_sized_to_their_content():
    doc = _table_doc(
        rows=(("H-300x300", 300, 10, 94.0, "합격"),),
        columns=("호칭", "외경", "두께", "중량(kg/m)", "판정"),
    )
    ws = _load(render_generic_xlsx(doc))["해석 결과"]

    # ⚠️ openpyxl 의 기본 열 너비는 None 이 아니라 13.0 이다. '>= 8' 같은 느슨한 임계값은
    #    고정 폭 시절에도 우연히 통과한다. 실제로 내용에 맞춰졌는지 값으로 못박는다.
    assert ws.column_dimensions["A"].width == _text_width("H-300x300") + 2   # 11
    assert ws.column_dimensions["D"].width == _text_width("중량(kg/m)") + 2  # 12 (한글 2폭)
    assert ws.column_dimensions["E"].width == 10  # 내용이 짧아 하한(_MIN_COL_WIDTH)에 걸린다


def test_undetermined_verdict_is_written_out_not_left_blank():
    """빈 칸은 '아직 안 채운 항목'처럼 보인다 — 판정 불가와 구분되어야 한다."""
    ws = _load(render_generic_xlsx(_doc(verdict=None)))["표지"]
    assert "판정 미확정" in [cell.value for cell in _cells(ws)]


def test_undetermined_verdict_is_marked_as_a_caution_not_a_pass():
    ws = _load(render_generic_xlsx(_doc(verdict=None)))["표지"]
    cell = next(c for c in _cells(ws) if c.value == "판정 미확정")
    assert cell.font.color.rgb.endswith("CC6600")


def test_first_table_header_row_is_frozen():
    """200행 넘는 표에서 스크롤하면 열 이름이 사라진다."""
    doc = _table_doc(rows=tuple((f"E{index}", "합격") for index in range(50)))
    ws = _load(render_generic_xlsx(doc))["해석 결과"]
    assert ws.freeze_panes is not None


@pytest.mark.parametrize(
    "word,colour",
    [("부적합", "CC0000"), ("failed", "CC0000"), ("주의", "CC6600"), ("적합", "1B7A3D")],
)
def test_renderer_highlights_every_word_the_adapter_understands(word, colour):
    """어댑터가 아는 낱말인데 렌더러가 모르면 실패 행이 조용히 강조를 잃는다."""
    ws = _load(render_generic_xlsx(_table_doc(rows=(("A", word),))))["해석 결과"]
    cell = next(c for c in _cells(ws) if c.value == word)
    assert cell.font.color.rgb.endswith(colour)


def test_an_unrelated_column_holding_a_verdict_word_does_not_paint_the_row():
    """무관한 열의 'OK' 가 규격 이탈 행을 합격 색으로 칠하던 문제.

    등록 어댑터가 없는 23개 App 이 전부 generic 표 경로를 쓰므로 실제로 도달한다.
    """
    doc = _table_doc(rows=(("W1", "OK", 8.4),), columns=("id", "inspector_note", "gap_mm"))
    ws = _load(render_generic_xlsx(doc))["해석 결과"]

    cell = next(c for c in _cells(ws) if c.value == "OK")
    assert cell.fill.patternType is None
    assert cell.font.color is None


def test_a_real_verdict_column_still_paints_the_row():
    doc = _table_doc(rows=(("W1", "불합격"),), columns=("id", "판정"))
    ws = _load(render_generic_xlsx(doc))["해석 결과"]

    cell = next(c for c in _cells(ws) if c.value == "불합격")
    assert cell.fill.patternType == "solid"


def test_renderer_and_adapter_share_one_verdict_vocabulary():
    """같은 낱말 목록을 두 곳에 적어 두면 언젠가 어긋난다 — 출처가 하나여야 한다."""
    from app.services.report import verdict_vocab
    from app.services.report.adapters import generic as adapter
    from app.services.report.renderers import generic_xlsx as renderer

    assert renderer._FAIL_WORDS is verdict_vocab.NEGATIVE_TOKENS
    assert renderer._WARN_WORDS is verdict_vocab.WARNING_TOKENS
    assert renderer._PASS_WORDS is verdict_vocab.POSITIVE_TOKENS
    assert adapter._VERDICT_NEGATIVE_TOKENS is verdict_vocab.NEGATIVE_TOKENS
    assert adapter._VERDICT_WARNING_TOKENS is verdict_vocab.WARNING_TOKENS
    assert adapter._VERDICT_POSITIVE_TOKENS is verdict_vocab.POSITIVE_TOKENS


def test_values_openpyxl_cannot_write_are_stringified_instead_of_crashing():
    """표 셀에 리스트·사전이 들어와도 리포트 생성이 죽지 않아야 한다.

    generic._table_from_rows 는 행 값에 리스트를 그대로 실을 수 있고, 그걸 openpyxl 에
    넘기면 저장 시 ValueError 로 리포트 전체가 실패한다. 계산서는 죽는 대신
    덜 예쁘게 나오는 쪽을 택한다.
    """
    doc = _table_doc(rows=((["a", "b"], {"k": 1}),), columns=("목록", "사전"))

    ws = _load(render_generic_xlsx(doc))["해석 결과"]

    values = [cell.value for cell in _cells(ws)]
    assert "['a', 'b']" in values
    assert "{'k': 1}" in values
