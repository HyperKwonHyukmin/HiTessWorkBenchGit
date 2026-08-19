"""리포트 오케스트레이터 — 레코드에서 bytes 까지."""
import io

import openpyxl
import pytest

from app import models
from app.services.report.service import (
    ReportNotAvailable,
    build_report_xlsx,
    report_capabilities,
)


def _record(**kwargs) -> models.Analysis:
    defaults = {
        "id": 7,
        "employee_id": "EMP001",
        "program_name": "Column Buckling Load Calculator",
        "project_name": "p",
        "status": "Success",
        "input_info": {"length_mm": 3000},
        "result_info": {"maxWorkingLoadTon": 12.5},
    }
    defaults.update(kwargs)
    return models.Analysis(**defaults)


def test_builds_xlsx_for_an_app_without_any_adapter(tmp_path):
    filename, data = build_report_xlsx(_record(), user_connection_base=str(tmp_path))

    assert filename.endswith(".xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames[0] == "표지"


def test_filename_carries_program_and_analysis_id(tmp_path):
    filename, _ = build_report_xlsx(_record(), user_connection_base=str(tmp_path))
    assert "7" in filename


def test_filename_names_the_app_in_words_a_person_reads(tmp_path):
    """program_id 슬러그는 다운로드 폴더에서 App 을 알아보는 데 도움이 안 된다.

    실제로 사용자는 받은 파일이 어느 App 결과인지 구분하지 못했다.
    """
    filename, _ = build_report_xlsx(_record(), user_connection_base=str(tmp_path))
    assert "Column_Buckling_Load_Calculator" in filename


def test_filename_carries_the_project_so_two_runs_differ(tmp_path):
    filename, _ = build_report_xlsx(
        _record(project_name="Deck A"), user_connection_base=str(tmp_path)
    )
    assert "Deck_A" in filename


def test_filename_drops_characters_that_are_illegal_in_a_path(tmp_path):
    filename, _ = build_report_xlsx(
        _record(project_name='a/b:c*d?"e<f>g|h'), user_connection_base=str(tmp_path)
    )
    assert not set(filename) & set(r'/\:*?"<>|')


def test_filename_stays_short_enough_for_windows(tmp_path):
    filename, _ = build_report_xlsx(
        _record(project_name="X" * 400), user_connection_base=str(tmp_path)
    )
    assert len(filename) <= 128


def test_rejects_a_record_without_results(tmp_path):
    with pytest.raises(ReportNotAvailable):
        build_report_xlsx(_record(result_info=None), user_connection_base=str(tmp_path))


def test_rejects_a_failed_record(tmp_path):
    with pytest.raises(ReportNotAvailable):
        build_report_xlsx(_record(status="Failed"), user_connection_base=str(tmp_path))


def test_unknown_program_name_still_produces_a_report(tmp_path):
    """미등록 App 도 파일명에 자기 이름을 남긴다.

    예전에는 program_id 가 없다는 이유로 'unknown' 을 적었는데, 그러면 서로 다른
    미등록 App 의 계산서가 다운로드 폴더에서 똑같아 보인다. 레지스트리에 없더라도
    레코드에 남은 이름이 사용자가 아는 유일한 단서다.
    """
    record = _record(program_name="아직 등록 안 된 App")
    filename, data = build_report_xlsx(record, user_connection_base=str(tmp_path))
    assert data
    assert "아직_등록_안_된_App" in filename


def test_applied_form_is_stated_on_the_cover_row_not_buried_in_notices(tmp_path):
    """서식 종류는 표지의 「적용 양식」 행이 말한다.

    유의 사항 블록은 이 해석 고유의 경고만 담는다 — 문서 메타 정보를 섞으면
    '왜 판정이 비었는가' 같은 정작 중요한 줄이 묻힌다.
    """
    _, data = build_report_xlsx(_record(), user_connection_base=str(tmp_path))
    wb = openpyxl.load_workbook(io.BytesIO(data))
    values = [cell.value for row in wb["표지"].iter_rows() for cell in row]

    assert "범용 서식" in values
    assert not any(isinstance(v, str) and "범용 서식으로 생성" in v for v in values)


def test_an_adapter_returning_the_wrong_type_falls_back_like_an_exception(tmp_path, monkeypatch):
    """예외만 막으면 폴백이 반쪽이다 — 조용한 None 반환도 같게 취급해야 한다."""
    from app.services.report import service as service_module

    monkeypatch.setattr(service_module, "get_adapter", lambda key: (lambda payload, meta: None))

    filename, data = build_report_xlsx(_record(), user_connection_base=str(tmp_path))

    assert data
    wb = openpyxl.load_workbook(io.BytesIO(data))
    values = [cell.value for row in wb["표지"].iter_rows() for cell in row]
    assert "전용 어댑터가 실패해 기본 서식으로 생성되었습니다." in values


def test_failed_provenance_lookup_is_distinguished_from_having_no_artifacts(tmp_path, monkeypatch):
    """'조회 실패'와 '원래 없음'은 다른 사실이다 — 같은 문구로 쓰면 승인자가 오해한다."""
    from app.services.report import service as service_module

    def _boom(*args, **kwargs):
        raise RuntimeError("passport 실패")

    monkeypatch.setattr(service_module, "build_analysis_passport", _boom)

    _, data = build_report_xlsx(_record(), user_connection_base=str(tmp_path))

    wb = openpyxl.load_workbook(io.BytesIO(data))
    values = [cell.value for sheet in wb for row in sheet.iter_rows() for cell in row]
    assert "조회 실패" in values
    assert "기록 없음" not in values
    assert any(isinstance(v, str) and "계보를 조회하지 못했습니다" in v for v in values)


def test_cover_and_verdict_sheet_use_the_same_wording(tmp_path):
    """같은 None 판정을 표지는 '판정 미확정', 판정 시트는 '판정 없음' 이라 부르던 문제.

    한 문서 안에서 같은 사실에 두 문구를 쓰면 승인자가 서로 다른 상태로 읽는다.
    단위 테스트가 두 곳을 따로만 봐서 실제 리포트를 열기 전까지 드러나지 않았다.
    """
    _, data = build_report_xlsx(_record(result_info={"note": "판정 키 없음"}),
                                user_connection_base=str(tmp_path))
    wb = openpyxl.load_workbook(io.BytesIO(data))

    cover = [c.value for row in wb["표지"].iter_rows() for c in row]
    verdict_sheet = [c.value for row in wb["판정"].iter_rows() for c in row]

    assert "판정 미확정" in cover
    assert "판정 미확정" in verdict_sheet
    assert "판정 없음" not in cover + verdict_sheet


def test_generic_path_explains_why_the_verdict_is_blank(tmp_path):
    """통과한 해석이 이유 없는 주황색 '미확정' 으로만 나오면 도구가 고장 난 것처럼 보인다."""
    _, data = build_report_xlsx(_record(result_info={"note": "판정 키 없음"}),
                                user_connection_base=str(tmp_path))
    wb = openpyxl.load_workbook(io.BytesIO(data))
    values = [c.value for row in wb["표지"].iter_rows() for c in row]

    assert any(isinstance(v, str) and "종합 판정을 표기하지 않습니다" in v for v in values)


def test_capabilities_lists_registered_programs():
    caps = report_capabilities()
    assert caps["truss-assessment"]["reportable"] is True
    assert caps["truss-assessment"]["hasTemplate"] is False
    assert "displayName" in caps["truss-assessment"]
