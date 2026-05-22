"""Daily/Weekly/Monthly 사용량 리포트 — 기간 계산·집계·Excel 빌더."""
import io
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from calendar import monthrange
from typing import Optional, Literal

PeriodType = Literal["daily", "weekly", "monthly"]
_WEEKDAY_KR = ["월", "화", "수", "목", "금", "토", "일"]


@dataclass(frozen=True)
class PeriodBounds:
    type: PeriodType
    start: datetime
    end: datetime
    prev_start: datetime
    prev_end: datetime
    label: str
    prev_label: str


def _end_of_day(d: date) -> datetime:
    return datetime(d.year, d.month, d.day, 23, 59, 59, 999999)


def _start_of_day(d: date) -> datetime:
    return datetime(d.year, d.month, d.day, 0, 0, 0)


def resolve_period(
    period: str,
    target: Optional[date],
    today: Optional[date] = None,
) -> PeriodBounds:
    """`period`가 가리키는 달력 기간의 경계를 반환.

    target이 None이면 '직전의 완료된 기간'을 기본값으로 사용한다.
    today는 테스트용 주입 포인트(없으면 실제 오늘).
    """
    if period not in ("daily", "weekly", "monthly"):
        raise ValueError(f"period는 daily, weekly, monthly 중 하나여야 합니다 (받은 값: {period!r})")

    today = today or date.today()

    # 기본값 보정
    if target is None:
        if period == "daily":
            target = today - timedelta(days=1)
        elif period == "weekly":
            target = today - timedelta(days=7)
        else:  # monthly
            target = today.replace(day=1) - timedelta(days=1)

    if target > today:
        raise ValueError("미래 기간은 조회할 수 없습니다.")

    if period == "daily":
        start_d = target
        end_d = target
        prev_start_d = target - timedelta(days=1)
        prev_end_d = prev_start_d
        weekday = _WEEKDAY_KR[start_d.weekday()]
        label = f"{start_d.isoformat()} ({weekday})"
        prev_label = prev_start_d.isoformat()

    elif period == "weekly":
        # Monday=0
        start_d = target - timedelta(days=target.weekday())
        end_d = start_d + timedelta(days=6)
        prev_start_d = start_d - timedelta(days=7)
        prev_end_d = end_d - timedelta(days=7)
        label = f"{start_d.isoformat()} ~ {end_d.isoformat()}"
        prev_label = f"{prev_start_d.isoformat()} ~ {prev_end_d.isoformat()}"

    else:  # monthly
        start_d = target.replace(day=1)
        last_day = monthrange(start_d.year, start_d.month)[1]
        end_d = start_d.replace(day=last_day)
        prev_end_d = start_d - timedelta(days=1)
        prev_start_d = prev_end_d.replace(day=1)
        label = f"{start_d.year}-{start_d.month:02d}"
        prev_label = f"{prev_start_d.year}-{prev_start_d.month:02d}"

    return PeriodBounds(
        type=period,
        start=_start_of_day(start_d),
        end=_end_of_day(end_d),
        prev_start=_start_of_day(prev_start_d),
        prev_end=_end_of_day(prev_end_d),
        label=label,
        prev_label=prev_label,
    )


from collections import Counter
from sqlalchemy.orm import Session
from app import models


def _count_new_users(db: Session, start: datetime, end: datetime, user_ids: set) -> int:
    """user_ids 중, 전체 이력상 첫 실행이 [start, end] 안에 있는 사용자 수."""
    if not user_ids:
        return 0
    from sqlalchemy import func as sqlfunc
    rows = (
        db.query(models.Analysis.employee_id, sqlfunc.min(models.Analysis.created_at))
          .filter(models.Analysis.employee_id.in_(user_ids))
          .group_by(models.Analysis.employee_id)
          .all()
    )
    return sum(1 for _, first_run in rows if start <= first_run <= end)


def aggregate_period(db: Session, period: str, start: datetime, end: datetime) -> dict:
    """기간 [start, end] 범위의 해석 데이터를 집계."""
    rows = (
        db.query(models.Analysis, models.User)
          .outerjoin(models.User, models.Analysis.employee_id == models.User.employee_id)
          .filter(models.Analysis.created_at >= start)
          .filter(models.Analysis.created_at <= end)
          .filter(models.Analysis.source != "WorkbenchSample")  # 샘플 실행은 통계 제외
          .all()
    )
    raw_rows = list(rows)
    stats_rows = [(a, u) for a, u in rows if not (u and u.is_developer)]
    total = len(stats_rows)

    if total == 0:
        return {
            "total": 0, "activePrograms": 0, "activeUsers": 0, "activeDepartments": 0,
            "avgPerDay": 0.0, "maxDay": 0,
            "busiestProgram": None, "peakHour": None, "newUsers": 0,
            "programs": [], "users": [], "departments": [],
            "timeBuckets": {"type": _bucket_type(period), "data": _empty_buckets(period, start, end)},
            "raw_rows": raw_rows,
        }

    program_map: dict = {}
    user_map: dict = {}
    dept_counter: Counter = Counter()
    day_counter: Counter = Counter()
    hour_counts = [0] * 24
    weekday_counts = [0] * 7

    for a, u in stats_rows:
        pname = a.program_name or "Unknown"
        eid = a.employee_id or "unknown"
        dept = (u.department if u and u.department else "Unknown")
        name = (u.name if u else "Deleted User")
        ts = a.created_at

        p = program_map.setdefault(pname, {"name": pname, "count": 0, "_users": set(), "lastRun": None})
        p["count"] += 1
        p["_users"].add(eid)
        if not p["lastRun"] or ts > p["lastRun"]:
            p["lastRun"] = ts

        u_row = user_map.setdefault(eid, {
            "employeeId": eid, "name": name, "department": dept,
            "count": 0, "_programs": set(), "lastRun": None,
        })
        u_row["count"] += 1
        u_row["_programs"].add(pname)
        if not u_row["lastRun"] or ts > u_row["lastRun"]:
            u_row["lastRun"] = ts

        dept_counter[dept] += 1
        day_counter[ts.date().isoformat()] += 1
        hour_counts[ts.hour] += 1
        weekday_counts[ts.weekday()] += 1

    programs = sorted(
        [
            {
                "name": p["name"],
                "count": p["count"],
                "share": round(p["count"] * 100 / total),
                "userCount": len(p["_users"]),
                "lastRun": p["lastRun"].isoformat() if p["lastRun"] else None,
            }
            for p in program_map.values()
        ],
        key=lambda r: r["count"], reverse=True,
    )
    users = sorted(
        [
            {
                "employeeId": u["employeeId"],
                "name": u["name"],
                "department": u["department"],
                "count": u["count"],
                "share": round(u["count"] * 100 / total),
                "programCount": len(u["_programs"]),
                "lastRun": u["lastRun"].isoformat() if u["lastRun"] else None,
            }
            for u in user_map.values()
        ],
        key=lambda r: r["count"], reverse=True,
    )

    covered_days = max(1, (end.date() - start.date()).days + 1)
    busiest_program = programs[0]["name"] if programs else None
    peak_hour_idx = max(range(24), key=lambda i: hour_counts[i]) if any(hour_counts) else None
    peak_hour = f"{peak_hour_idx:02d}시" if peak_hour_idx is not None else None

    time_buckets = _build_time_buckets(period, start, end, hour_counts, weekday_counts, day_counter)

    return {
        "total": total,
        "activePrograms": len(program_map),
        "activeUsers": len(user_map),
        "activeDepartments": len(dept_counter),
        "avgPerDay": round(total / covered_days, 1),
        "maxDay": max(day_counter.values()) if day_counter else 0,
        "busiestProgram": busiest_program,
        "peakHour": peak_hour,
        "newUsers": _count_new_users(db, start, end, set(user_map.keys())),
        "programs": programs,
        "users": users,
        "departments": [{"name": k, "count": v} for k, v in dept_counter.most_common()],
        "timeBuckets": time_buckets,
        "raw_rows": raw_rows,
    }


def _bucket_type(period: str) -> str:
    return {"daily": "hour", "weekly": "weekday", "monthly": "dayOfMonth"}[period]


def _empty_buckets(period: str, start: datetime, end: datetime) -> list:
    if period == "daily":
        return [{"label": f"{h:02d}시", "count": 0} for h in range(24)]
    if period == "weekly":
        return [{"label": _WEEKDAY_KR[i], "count": 0} for i in range(7)]
    days = (end.date() - start.date()).days + 1
    return [{"label": str(d), "count": 0} for d in range(1, days + 1)]


def _build_time_buckets(period, start, end, hour_counts, weekday_counts, day_counter) -> dict:
    if period == "daily":
        data = [{"label": f"{h:02d}시", "count": hour_counts[h]} for h in range(24)]
    elif period == "weekly":
        data = [{"label": _WEEKDAY_KR[i], "count": weekday_counts[i]} for i in range(7)]
    else:
        days = (end.date() - start.date()).days + 1
        data = []
        for offset in range(days):
            d = (start.date() + timedelta(days=offset))
            data.append({"label": str(d.day), "count": day_counter.get(d.isoformat(), 0)})
    return {"type": _bucket_type(period), "data": data}


def compute_deltas(current: dict, previous: dict) -> dict:
    """current/previous summary 딕셔너리의 주요 필드 차이를 계산."""
    keys = ("total", "activeUsers", "activePrograms", "avgPerDay")
    out = {}
    for k in keys:
        cur = current.get(k, 0)
        prev = previous.get(k, 0)
        abs_delta = round(cur - prev, 1) if isinstance(cur, float) or isinstance(prev, float) else cur - prev
        if prev == 0:
            pct = None
        else:
            pct = round((cur - prev) * 100 / prev, 1)
        out[k] = {"abs": abs_delta, "pct": pct}
    return out


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="FF2563EB")


def _apply_header_style(cell):
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = Alignment(horizontal="center")


def _fmt_delta(d) -> str:
    if not d:
        return ""
    sign = "+" if d["abs"] >= 0 else ""
    pct = "" if d["pct"] is None else f" ({sign}{d['pct']}%)"
    return f"{sign}{d['abs']}{pct}"


def build_report_xlsx(bounds: PeriodBounds, current: dict, previous: dict, deltas: dict):
    """6개 시트의 Excel 리포트를 BytesIO로 반환."""
    import io as _bio
    wb = Workbook()
    _build_summary_sheet(wb, bounds, current, previous, deltas)
    _build_table_sheet(wb, "Programs", current["programs"],
                       columns=[("순위", None), ("프로그램", "name"), ("실행수", "count"),
                                ("점유율(%)", "share"), ("사용자수", "userCount"), ("최근실행", "lastRun")])
    _build_table_sheet(wb, "Users", current["users"],
                       columns=[("순위", None), ("사번", "employeeId"), ("이름", "name"),
                                ("부서", "department"), ("실행수", "count"), ("점유율(%)", "share"),
                                ("사용프로그램수", "programCount"), ("최근실행", "lastRun")])
    _build_dept_sheet(wb, current)
    _build_time_sheet(wb, current["timeBuckets"])
    _build_raw_sheet(wb, current["raw_rows"])

    buf = _bio.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_summary_sheet(wb, bounds, current, previous, deltas):
    ws = wb.active
    ws.title = "Summary"
    ws.append(["항목", "값", "전 기간 대비"])
    for c in ws[1]:
        _apply_header_style(c)
    rows = [
        ("기간", bounds.label, ""),
        ("이전 기간", bounds.prev_label, ""),
        ("총 실행 수", current["total"], _fmt_delta(deltas.get("total"))),
        ("활성 프로그램 수", current["activePrograms"], _fmt_delta(deltas.get("activePrograms"))),
        ("활성 사용자 수", current["activeUsers"], _fmt_delta(deltas.get("activeUsers"))),
        ("활성 부서 수", current["activeDepartments"], ""),
        ("일평균 실행 수", current["avgPerDay"], _fmt_delta(deltas.get("avgPerDay"))),
        ("최대 일일 실행 수", current["maxDay"], ""),
        ("최다 사용 프로그램", current["busiestProgram"] or "-", ""),
        ("피크 시간대", current["peakHour"] or "-", ""),
        ("신규 사용자 수", current["newUsers"], ""),
    ]
    for r in rows:
        ws.append(list(r))
    for col in (1, 2, 3):
        ws.column_dimensions[chr(64 + col)].width = 24


def _build_table_sheet(wb, title, rows, columns):
    ws = wb.create_sheet(title)
    headers = [c[0] for c in columns]
    ws.append(headers)
    for c in ws[1]:
        _apply_header_style(c)
    for i, r in enumerate(rows, start=1):
        row_values = []
        for label, key in columns:
            if key is None:
                row_values.append(i)
            else:
                row_values.append(r.get(key, ""))
        ws.append(row_values)
    for idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + idx)].width = 18


def _build_dept_sheet(wb, current):
    ws = wb.create_sheet("Departments")
    ws.append(["부서", "실행수", "점유율(%)"])
    for c in ws[1]:
        _apply_header_style(c)
    total = current["total"] or 1
    for d in current["departments"]:
        ws.append([d["name"], d["count"], round(d["count"] * 100 / total)])
    for idx in range(1, 4):
        ws.column_dimensions[chr(64 + idx)].width = 20


def _build_time_sheet(wb, time_buckets):
    ws = wb.create_sheet("Time Distribution")
    label_header = {"hour": "시간대", "weekday": "요일", "dayOfMonth": "일자"}[time_buckets["type"]]
    ws.append([label_header, "실행수"])
    for c in ws[1]:
        _apply_header_style(c)
    for b in time_buckets["data"]:
        ws.append([b["label"], b["count"]])
    for idx in range(1, 3):
        ws.column_dimensions[chr(64 + idx)].width = 18


def _build_raw_sheet(wb, raw_rows):
    ws = wb.create_sheet("Raw Data")
    ws.append(["ID", "프로젝트", "프로그램", "사번", "이름", "부서", "상태", "실행시각", "개발자"])
    for c in ws[1]:
        _apply_header_style(c)
    for a, u in raw_rows:
        ws.append([
            a.id,
            getattr(a, "project_name", "") or "",
            a.program_name or "",
            a.employee_id or "",
            (u.name if u else "Deleted User"),
            (u.department if u and u.department else "Unknown"),
            a.status or "",
            a.created_at.isoformat() if a.created_at else "",
            "Y" if (u and u.is_developer) else "",
        ])
    widths = [8, 24, 28, 12, 14, 18, 12, 22, 8]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w
