"""ReportDoc → XLSX bytes (양식 없는 범용 서식).

디스크에 쓰지 않는다 — 저장하는 순간 사내 DRM 이 암호화한다. BytesIO 만 쓴다.
색 팔레트는 assessment_service._json_to_xlsx_bytes 와 맞춘다(PRODUCT.md Trust Blue).
"""
from __future__ import annotations

import io
from datetime import date, datetime, time
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .. import verdict_vocab
from ..models import ReportDoc, ReportSection

_HDR_FILL = PatternFill("solid", fgColor="002554")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=9)
_SEC_FILL = PatternFill("solid", fgColor="D6E0F0")
_SEC_FONT = Font(bold=True, color="002554", size=10)
_TITLE_FONT = Font(bold=True, color="002554", size=14)
_BASE_FONT = Font(size=9)
_LABEL_FONT = Font(bold=True, size=9)
_LEFT = Alignment(horizontal="left", vertical="center")
_CENTER = Alignment(horizontal="center", vertical="center")

# 판정 강조 — 색만으로 알리지 않는다(PRODUCT.md 접근성). 글자는 이미 '불합격'/'합격'이고
# 여기에 색을 덧입혀, 긴 표를 훑을 때 실패 행이 눈에 걸리게 한다.
# assessment_service._json_to_xlsx_bytes 와 같은 팔레트를 쓴다.
_FAIL_FILL = PatternFill("solid", fgColor="FFE4E4")
_FAIL_FONT = Font(bold=True, color="CC0000", size=9)
_WARN_FONT = Font(bold=True, color="CC6600", size=9)
_PASS_FONT = Font(bold=True, color="1B7A3D", size=9)

# ⚠️ 낱말 목록을 여기에 따로 적지 않는다. 어댑터가 '부적합'을 불합격으로 읽는데
#    렌더러가 그 행을 강조하지 않으면, '실패 행이 눈에 걸려야 한다'는 약속이
#    어휘 불일치만으로 조용히 깨진다. 단일 출처에서 가져온다.
_FAIL_WORDS = verdict_vocab.NEGATIVE_TOKENS
_WARN_WORDS = verdict_vocab.WARNING_TOKENS
_PASS_WORDS = verdict_vocab.POSITIVE_TOKENS
_STATUS_FONT = {"fail": _FAIL_FONT, "warn": _WARN_FONT, "pass": _PASS_FONT}

# ⚠️ 판정어를 행 전체에서 찾지 않는다. 표 열은 그냥 라벨이라 어떤 열이 판정인지 표시가 없고,
#    무관한 열('inspector_note' 값이 우연히 "OK')이 행 전체를 합격 색으로 칠해 버린다.
#    실제로 규격을 벗어난 행이 초록으로, 문제 있는 행이 무색으로 나오는 역전이 생긴다.
#    adapters/generic._VERDICT_KEYS 가 최상위 키를 제한하는 것과 같은 취지를 열 이름에 적용한다.
_VERDICT_COLUMN_NAMES: frozenset[str] = frozenset({
    "result", "verdict", "judgement", "judgment", "result_status",
    "assessment", "판정", "결과",
})

_SHEET_NAME_LIMIT = 31
# Excel 시트 이름에 쓸 수 없는 문자.
_FORBIDDEN = ':\\/?*[]'

# 열 너비 자동 맞춤 범위. 표 열이 3개를 넘으면 고정 너비로는 헤더가 잘린다.
_MIN_COL_WIDTH = 10
_MAX_COL_WIDTH = 52


def _safe_sheet_name(title: str, used: set[str]) -> str:
    cleaned = "".join("_" if ch in _FORBIDDEN else ch for ch in (title or "섹션")).strip() or "섹션"
    cleaned = cleaned[:_SHEET_NAME_LIMIT]
    candidate = cleaned
    suffix = 2
    while candidate.casefold() in used:
        tail = f"_{suffix}"
        candidate = cleaned[: _SHEET_NAME_LIMIT - len(tail)] + tail
        suffix += 1
    used.add(candidate.casefold())
    return candidate


def _cell_value(value):
    """openpyxl 이 셀에 쓸 수 없는 값은 텍스트로 굳힌다.

    generic._table_from_rows 는 행 값에 리스트·사전을 그대로 실을 수 있다. 그대로
    넘기면 wb.save 에서 ValueError 가 나 리포트 생성 전체가 죽는다 — 계산서는
    죽는 대신 덜 예쁘게 나오는 쪽이 낫다.
    """
    if value is None or isinstance(value, (str, int, float, bool, Decimal, datetime, date, time)):
        return value
    return str(value)


def _text_width(value) -> int:
    """대략적인 표시 폭. 한글·한자는 라틴 문자의 두 배로 친다."""
    text = "" if value is None else str(value)
    return sum(2 if ord(ch) > 0x2E80 else 1 for ch in text)


def _autofit(ws, widths: dict[int, int]) -> None:
    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = max(
            _MIN_COL_WIDTH, min(width + 2, _MAX_COL_WIDTH)
        )


def _status_of(value) -> str | None:
    """셀 값이 판정어인지. 표 안에 묻힌 실패를 찾아내는 데 쓴다."""
    if not isinstance(value, str):
        return None
    token = value.strip().casefold()
    if token in _FAIL_WORDS:
        return "fail"
    if token in _WARN_WORDS:
        return "warn"
    if token in _PASS_WORDS:
        return "pass"
    return None


def _row_status(row, columns: tuple[str, ...]) -> str | None:
    """행의 판정. **판정 열에서만** 읽는다. 실패가 하나라도 있으면 실패로 본다.

    판정으로 볼 만한 열이 하나도 없으면 아무 강조도 하지 않는다 — 잘못된 강조는
    강조가 없는 것보다 나쁘다. 결재자의 눈을 엉뚱한 행으로 끌기 때문이다.
    """
    seen: set[str] = {
        status
        for name, value in zip(columns, row)
        if str(name).strip().casefold() in _VERDICT_COLUMN_NAMES
        and (status := _status_of(value))
    }
    for level in ("fail", "warn", "pass"):
        if level in seen:
            return level
    return None


def _verdict_cell(verdict: str | None):
    """판정 칸의 표기와 글꼴.

    비어 있으면 '확인 필요'라고 분명히 적는다 — 빈 칸은 '아직 안 채운 항목'처럼
    보여서 '판정할 근거가 없다'와 구분되지 않는다. 어댑터가 근거 부족을 이유로
    None 을 낸 노력이 렌더 층에서 도로 사라지면 안 된다.

    문구는 서비스(판정 시트)와 같은 상수를 쓴다 — 두 곳에 따로 적어 두면
    표지와 판정 시트가 같은 사실을 다른 낱말로 말하게 된다(실제로 그랬다).
    """
    if verdict is None:
        return verdict_vocab.UNDETERMINED_VERDICT, _WARN_FONT
    return verdict, _STATUS_FONT.get(_status_of(verdict), _BASE_FONT)


def _write_cover(ws, doc: ReportDoc) -> None:
    ws["A1"] = f"{doc.meta.display_name} 해석 계산서"
    ws["A1"].font = _TITLE_FONT
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 46

    verdict_text, verdict_font = _verdict_cell(doc.verdict)
    # 판정 개념이 없는 App(허용하중·단면 특성 산출 등)에는 판정 행을 아예 두지 않는다.
    # 빈 판정 칸은 '아직 안 나왔다'로 읽혀서, 판정할 것이 없다는 사실과 구분되지 않는다.
    show_verdict = doc.meta.verdict_kind != "none"
    rows = [
        ("해석 App", doc.meta.display_name),
        ("프로젝트", doc.meta.project_name),
        ("수행자 사번", doc.meta.employee_id),
        ("수행 일시", doc.meta.created_at.strftime("%Y-%m-%d %H:%M") if doc.meta.created_at else None),
        ("작업 상태", doc.meta.status),
        *([("판정", verdict_text)] if show_verdict else []),
        ("적용 양식", "사내 표준 양식" if doc.template_applied else "범용 서식"),
    ]
    row_ptr = 3
    for label, value in rows:
        ws.cell(row=row_ptr, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=row_ptr, column=1).alignment = _LEFT
        cell = ws.cell(row=row_ptr, column=2, value=value)
        # 판정 줄만 색을 덧입힌다 — 글자('불합격')는 그대로 두므로 색에만 기대지 않는다.
        cell.font = verdict_font if label == "판정" else _BASE_FONT
        cell.alignment = _LEFT
        row_ptr += 1

    if doc.notices:
        row_ptr += 1
        header = ws.cell(row=row_ptr, column=1, value="유의 사항")
        header.fill = _SEC_FILL
        header.font = _SEC_FONT
        row_ptr += 1
        for notice in doc.notices:
            ws.cell(row=row_ptr, column=1, value=notice).font = _BASE_FONT
            row_ptr += 1


def _write_section(ws, section: ReportSection) -> None:
    widths: dict[int, int] = {}
    freeze_at: int | None = None  # 첫 표의 헤더 아래 — 스크롤해도 열 이름이 남게

    def put(row, column, value, *, font=_BASE_FONT, fill=None, align=None):
        cell = ws.cell(row=row, column=column, value=_cell_value(value))
        cell.font = font
        if fill is not None:
            cell.fill = fill
        if align is not None:
            cell.alignment = align
        widths[column] = max(widths.get(column, 0), _text_width(value))
        return cell

    row_ptr = 1
    put(row_ptr, 1, section.title, font=_SEC_FONT, fill=_SEC_FILL, align=_LEFT)
    row_ptr += 2

    for item in section.fields:
        put(row_ptr, 1, item.label, font=_LABEL_FONT)
        put(row_ptr, 2, item.value)
        if item.unit:
            put(row_ptr, 3, item.unit)
        if item.note:
            put(row_ptr, 4, item.note)
        row_ptr += 1

    for table in section.tables:
        row_ptr += 1
        put(row_ptr, 1, table.title, font=_SEC_FONT)
        row_ptr += 1
        for col_index, column in enumerate(table.columns, start=1):
            put(row_ptr, col_index, column, font=_HDR_FONT, fill=_HDR_FILL, align=_CENTER)
        row_ptr += 1
        if freeze_at is None:
            freeze_at = row_ptr
        for row in table.rows:
            # 긴 표에 묻힌 실패 행은 훑어서는 안 보인다. 글자는 그대로 두고 색을 덧입힌다.
            status = _row_status(row, table.columns)
            font = _STATUS_FONT.get(status, _BASE_FONT)
            fill = _FAIL_FILL if status == "fail" else None
            for col_index, value in enumerate(row, start=1):
                put(row_ptr, col_index, value, font=font, fill=fill)
            row_ptr += 1
        if table.note:
            put(row_ptr, 1, table.note)
            row_ptr += 1

    _autofit(ws, widths)
    if freeze_at is not None:
        # 200행 넘는 표에서 스크롤하면 열 이름이 사라진다. 첫 표 헤더를 고정한다.
        ws.freeze_panes = f"A{freeze_at}"


def render_generic_xlsx(doc: ReportDoc) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    used: set[str] = set()
    cover = wb.create_sheet(_safe_sheet_name("표지", used))
    _write_cover(cover, doc)

    for section in doc.ordered_sections():
        ws = wb.create_sheet(_safe_sheet_name(section.title, used))
        _write_section(ws, section)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
