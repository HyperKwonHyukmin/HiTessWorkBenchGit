"""Truss Structural Assessment 해석 백그라운드 실행 로직."""
import io
import os
import csv
import json
import logging
from datetime import datetime

from .analysis_runner import (
    get_backend_dir,
    mark_complete,
    mark_running,
    record_analysis,
    run_engine,
    update_progress,
)

logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────
# JSON → XLSX 메모리 변환 (DRM 우회: 디스크에 저장하지 않음)
# ──────────────────────────────────────────────────────────
def _json_to_xlsx_bytes(json_path: str) -> bytes:
    """
    TrussAssessment JSON 결과를 openpyxl로 XLSX bytes로 변환합니다.
    디스크에 파일을 쓰지 않고 메모리(BytesIO)에서만 처리하므로
    회사 DRM 소프트웨어의 자동 암호화를 피할 수 있습니다.

    시트 구성: Load Case별 1개 시트
      - SUMMARY (Set별 최대값)
      - ELEMENT ASSESSMENT (전체 부재)
      - DISTRIBUTION PANEL (하중분산판, 데이터 있을 때만)
      - SIDE SUPPORT (데이터 있을 때만)
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    with open(json_path, encoding='utf-8-sig') as f:
        data = json.load(f)

    load_cases = data.get('loadCases', [])
    wb = Workbook()
    wb.remove(wb.active)  # 기본 빈 시트 제거

    # ── 스타일 ──
    HDR_FILL  = PatternFill('solid', fgColor='002554')   # 헤더 배경 (네이비)
    HDR_FONT  = Font(bold=True, color='FFFFFF', size=9)
    SEC_FILL  = PatternFill('solid', fgColor='D6E0F0')   # 섹션 타이틀 배경
    SEC_FONT  = Font(bold=True, color='002554', size=10)
    FAIL_FILL = PatternFill('solid', fgColor='FFE4E4')   # FAIL 행 배경
    FAIL_FONT = Font(bold=True, color='CC0000', size=9)
    WARN_FONT = Font(bold=True, color='CC6600', size=9)  # assessment 0.8~1.0
    BASE_FONT = Font(size=9)
    CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=False)

    COL_LABELS = {
        'element': 'Element',      'set': 'Set',            'property': 'Property',
        'axial': 'Axial',          'bending': 'Bending',
        'allowAxial': 'Allow Axial', 'allowBending': 'Allow Bending',
        'assessment': 'Assessment', 'result': 'Result',
        'leg': 'Leg',              'condition': 'Condition',
        'reactionForce': 'Reaction Force',
        'allowBF03': 'Allow BF-03', 'allowBF02': 'Allow BF-02', 'allowBF06': 'Allow BF-06',
        'panel': 'Panel Type',
        'support': 'Support Node', 'reaction': 'Reaction',
        'loadCaseId': 'LC',
    }

    def write_table(ws, row_ptr, section_title, rows):
        if not rows:
            return row_ptr
        headers = list(rows[0].keys())

        # 섹션 타이틀
        c = ws.cell(row=row_ptr, column=1, value=section_title)
        c.fill = SEC_FILL
        c.font = SEC_FONT
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=row_ptr, start_column=1,
                       end_row=row_ptr, end_column=max(len(headers), 1))
        row_ptr += 1

        # 헤더 행
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row_ptr, column=ci, value=COL_LABELS.get(h, h))
            c.fill = HDR_FILL
            c.font = HDR_FONT
            c.alignment = CENTER
        row_ptr += 1

        # 데이터 행
        for row in rows:
            is_fail = str(row.get('result', '')).upper() == 'FAIL'
            for ci, h in enumerate(headers, 1):
                val = row.get(h)
                # 정수 컬럼은 int 변환, 나머지 float는 소수점 4자리
                if isinstance(val, float):
                    if h in ('element', 'set', 'property', 'leg', 'support', 'loadCaseId'):
                        val = int(round(val))
                    else:
                        val = round(val, 4)

                c = ws.cell(row=row_ptr, column=ci, value=val)
                c.font = BASE_FONT
                c.alignment = CENTER

                if is_fail:
                    c.fill = FAIL_FILL
                if h == 'result' and is_fail:
                    c.font = FAIL_FONT
                elif h == 'assessment' and val is not None:
                    n = float(val)
                    if n >= 1.0:
                        c.font = FAIL_FONT
                    elif n >= 0.8:
                        c.font = WARN_FONT
            row_ptr += 1

        return row_ptr + 1  # 섹션 사이 빈 행

    # LC별, 섹션별로 각각 독립 시트 생성
    # 시트명 예: "LC1 Summary", "LC1 Element Assessment", "LC1 Distribution Panel", "LC1 Side Support"
    for lc in load_cases:
        lc_id = lc.get('loadCaseIndex', '?')

        sections = [
            (f'LC{lc_id} Summary',              'SUMMARY  (Set별 최대값)',         lc.get('summary', [])),
            (f'LC{lc_id} Element Assessment',   'ELEMENT ASSESSMENT  (전체 부재)', lc.get('elementAssessment', [])),
            (f'LC{lc_id} Distribution Panel',   'DISTRIBUTION PANEL  (하중분산판)',lc.get('distributionPanel', [])),
            (f'LC{lc_id} Side Support',         'SIDE SUPPORT',                    lc.get('sideSupport', [])),
        ]

        for sheet_title, section_title, rows in sections:
            if not rows:
                continue
            ws = wb.create_sheet(title=sheet_title[:31])  # Excel 시트명 31자 제한
            write_table(ws, 1, f'■ {section_title}', rows)

            # 컬럼 너비: 상위 5개 셀 기준 (대용량 데이터 성능 보호)
            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                max_len = max((len(str(c.value)) for c in list(col)[:5] if c.value), default=8)
                ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _json_to_csv(json_path: str, work_dir: str, base_name: str) -> dict:
    """
    TrussAssessment JSON 결과를 CSV 파일 2종으로 변환합니다.
    - {base_name}_Summary.csv        : 각 하중 케이스별 세트 대표값 (summary)
    - {base_name}_ElementResult.csv  : 전체 부재 평가 결과 (elementAssessment)
    DRM 정책은 .xlsx에만 적용되므로 CSV는 그대로 다운로드 가능합니다.
    반환값: { "CSV_Summary": path, "CSV_ElementResult": path }
    """
    generated = {}
    try:
        with open(json_path, encoding="utf-8") as f:
            data = json.load(f)

        load_cases = data.get("loadCases", [])
        if not load_cases:
            return generated

        # ── Summary CSV ──────────────────────────────────
        summary_rows = []
        for lc in load_cases:
            lc_id = lc.get("loadCaseIndex", "")
            for row in lc.get("summary", []):
                summary_rows.append({"loadCaseId": lc_id, **row})

        if summary_rows:
            summary_path = os.path.join(work_dir, f"{base_name}_Summary.csv")
            headers = list(summary_rows[0].keys())
            with open(summary_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(summary_rows)
            generated["CSV_Summary"] = summary_path

        # ── ElementResult CSV ─────────────────────────────
        elem_rows = []
        for lc in load_cases:
            lc_id = lc.get("loadCaseIndex", "")
            for row in lc.get("elementAssessment", []):
                elem_rows.append({"loadCaseId": lc_id, **row})

        if elem_rows:
            elem_path = os.path.join(work_dir, f"{base_name}_ElementResult.csv")
            headers = list(elem_rows[0].keys())
            with open(elem_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(elem_rows)
            generated["CSV_ElementResult"] = elem_path

        # ── DistributionPanel CSV ─────────────────────────
        panel_rows = []
        for lc in load_cases:
            lc_id = lc.get("loadCaseIndex", "")
            for row in lc.get("distributionPanel", []):
                panel_rows.append({"loadCaseId": lc_id, **row})

        if panel_rows:
            panel_path = os.path.join(work_dir, f"{base_name}_DistributionPanel.csv")
            headers = list(panel_rows[0].keys())
            with open(panel_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(panel_rows)
            generated["CSV_DistributionPanel"] = panel_path

        # ── SideSupport CSV ───────────────────────────────
        support_rows = []
        for lc in load_cases:
            lc_id = lc.get("loadCaseIndex", "")
            for row in lc.get("sideSupport", []):
                support_rows.append({"loadCaseId": lc_id, **row})

        if support_rows:
            support_path = os.path.join(work_dir, f"{base_name}_SideSupport.csv")
            headers = list(support_rows[0].keys())
            with open(support_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(support_rows)
            generated["CSV_SideSupport"] = support_path

    except Exception as e:
        # CSV 생성 실패는 전체 해석을 막지 않음
        generated["CSV_Error"] = str(e)

    return generated


def task_execute_assessment(job_id: str, bdf_path: str, work_dir: str, employee_id: str, timestamp: str, source: str):
  """
  TrussAssessment.exe 엔진을 호출하고, 결과로 생성된 json 파일을 스캔하여 반환합니다.
  xlsx는 회사 DRM에 의해 즉시 암호화되므로 JSON → CSV 변환 결과를 함께 제공합니다.
  """
  mark_running(job_id, "Initiating Assessment Solver...", progress=10)

  status_msg = "Success"
  engine_output = ""
  result_data = {}

  exe_dir = os.path.join(get_backend_dir(), "InHouseProgram", "TrussAssessment")
  exe_path = os.path.join(exe_dir, "TrussAssessment.exe")

  try:
    if not os.path.exists(exe_path):
      raise FileNotFoundError(f"Executable not found: {exe_path}")

    update_progress(job_id, 40, "Running Nastran Analysis & Evaluation...")

    # 첫 번째 인자로 bdf_path 전달 (작업 폴더 기준 실행)
    status_msg, engine_output = run_engine(
      [exe_path, bdf_path], work_dir, timeout=600, engine_label="TrussAssessment",
    )

    if status_msg == "Success":
      update_progress(job_id, 80, "Extracting Results & Converting to CSV...")

      # 결과 파일 스캔 (다중 Case 및 대소문자 확장자 완벽 대응)
      bdf_stem = os.path.splitext(os.path.basename(bdf_path))[0]
      json_count = 0
      for f in os.listdir(work_dir):
        full_path = os.path.join(work_dir, f)
        lower_f = f.lower()

        # xlsx: DRM 정책으로 암호화되므로 경로만 기록 (실제 다운로드는 CSV 사용 권장)
        if lower_f.endswith('.xlsx') and not f.startswith('~'):
          name_without_ext = os.path.splitext(f)[0]
          result_data[f"Excel_{name_without_ext}"] = full_path

        elif lower_f.endswith('.json'):
          json_count += 1
          name_without_ext = os.path.splitext(f)[0]
          result_data[f"JSON_{name_without_ext}"] = full_path

          # JSON → CSV 변환 (DRM 우회)
          update_progress(job_id, 80, f"Converting {f} to CSV...")
          csv_files = _json_to_csv(full_path, work_dir, name_without_ext)
          result_data.update(csv_files)

        elif lower_f.endswith('.f06'):
          target_name = f"{bdf_stem}.f06"
          target_path = os.path.join(work_dir, target_name)
          if f != target_name:
            os.rename(full_path, target_path)
            full_path = target_path
          result_data[f"F06_{bdf_stem}"] = full_path

        elif lower_f.endswith('.op2'):
          target_name = f"{bdf_stem}.op2"
          target_path = os.path.join(work_dir, target_name)
          if f != target_name:
            os.rename(full_path, target_path)
            full_path = target_path
          result_data[f"OP2_{bdf_stem}"] = full_path

      # BDF 원본도 결과로 함께 반환
      result_data["bdf"] = bdf_path

      if json_count == 0:
        engine_output += "\n[Warning] JSON result files were NOT found in the user's work directory. C# 엔진의 출력 경로를 확인하세요."
  except Exception as e:
    # exe_path 미존재(FileNotFoundError), 결과 파일 스캔 중 예기치 못한 오류 등을 일괄 처리.
    # subprocess 자체 오류는 run_engine 내부에서 처리되어 (Failed, 메시지)로 이미 반환됨.
    status_msg = "Failed"
    logger.error("TrussAssessment unexpected error: %s", str(e), exc_info=True)
    engine_output = "예기치 않은 오류가 발생했습니다. 관리자에게 문의하세요."

  update_progress(job_id, 95, "Saving to Database...")

  date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
  project_data, db_err = record_analysis(
    project_name=f"Truss Assessment_{date_str}",
    program_name="Truss Assessment",
    employee_id=employee_id,
    status=status_msg,
    input_info={"bdf_model": bdf_path},
    result_info=result_data if status_msg == "Success" else None,
    source=source,
  )
  if db_err is not None:
    status_msg = "Failed"
    engine_output += f"\nDB Error: {db_err}"

  mark_complete(job_id, status_msg, engine_output, project_data)
