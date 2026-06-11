"""Carling 리포트 생성 — openpyxl 인메모리 렌더링.

리포트 템플릿은 HHI DRM '비적용' 위치(사내 스토리지 공유 경로)에 평문(.xlsm=정상 ZIP)으로
보관한다. 평문 템플릿은 openpyxl 로 직접 열 수 있으므로, Excel COM/DRM/디스크 저장 없이
메모리(BytesIO)에서 값만 채워 .xlsx bytes 로 반환한다(TrussAssessment export-xlsx 와 동일 전략).

이 방식의 이점:
  - DRM 무관: 디스크에 저장하지 않아 DRM 자동 암호화 자체가 발생하지 않는다.
  - Excel 불필요: 서버에 Excel/DRM 에이전트가 없어도 동작한다(COM 의존 제거).
  - 경로 제약 회피: Excel 이 거부하는 '[' ']' 가 포함된 공유 경로도 openpyxl 은 그대로 읽는다.

템플릿 소스 디렉터리는 env(CARLING_REPORT_DIR) > 공유 스토리지 > 로컬 순으로 결정한다
(로컬 폴백 템플릿은 DRM 암호화 상태이면 openpyxl 로 못 열 수 있음 — 공유 경로 우선).
"""
import io
import logging
import os
import re
from datetime import datetime

import openpyxl
from fastapi import HTTPException

logger = logging.getLogger(__name__)

_SERVICES_DIR = os.path.dirname(os.path.abspath(__file__))
_APP_DIR = os.path.dirname(_SERVICES_DIR)
_BACKEND_DIR = os.path.dirname(_APP_DIR)

# 템플릿 소스 디렉터리 후보.
# 1순위: 사내 스토리지 공유 경로(DRM 미적용 — 평문 템플릿 보관). env 로 오버라이드 가능.
# 폴백: 로컬 InHouseProgram(템플릿이 DRM 암호화 상태이면 openpyxl 로딩 실패 가능).
_SHARE_REPORT_DIR = r"\\storage.hpc.hd.com\a476854\00_PROJECT\AA_300_CF44\[개인 자료]\권혁민 책임연구원\HiTessWorkBench\Reports\CarlingCalculator"
_LOCAL_REPORT_DIR = os.path.join(_BACKEND_DIR, "InHouseProgram", "CarlingCalculator", "Report")

# mode → load_type → 템플릿 파일명
_TEMPLATES = {
    "free": {
        "concentrated": "Carling Free Calculator Report_Concentrated.xlsm",
        "distributed": "Carling Free Calculator Report_Distributed.xlsm",
    },
    "optimization": {
        "concentrated": "Carling Design Optimization_Concentrated.xlsm",
        "distributed": "Carling Design Optimization_Distributed.xlsm",
    },
}

# mode → 파일명 라벨
_MODE_LABEL = {"free": "Free", "optimization": "Optimization"}

# 외부 워크북 참조 IF 수식: =IF([1]Opt.!E4="집중 하중", "<concentrated>", "<distributed>")
_EXT_IF_RE = re.compile(r'=IF\([^,]*,\s*"([^"]*)"\s*,\s*"([^"]*)"\s*\)')


def _resolve_report_dir() -> str:
    """템플릿 소스 디렉터리 결정: env(CARLING_REPORT_DIR) > 공유 스토리지 > 로컬."""
    for cand in (os.environ.get("CARLING_REPORT_DIR"), _SHARE_REPORT_DIR, _LOCAL_REPORT_DIR):
        if cand and os.path.isdir(cand):
            return cand
    return _LOCAL_REPORT_DIR


def _resolve_external_formulas(ws, is_concentrated: bool) -> None:
    """외부 워크북 참조([1]Opt.!…)가 든 수식을 load_type 기준 리터럴로 치환한다.

    openpyxl 재저장 시 외부참조가 #REF! 로 깨지므로, 단위/라벨 수식을 미리 평문으로 굳힌다.
    매칭 안 되는 외부참조 수식은 공란 처리해 깨진 참조가 남지 않게 한다.
    """
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("=") and "[" in v:
                m = _EXT_IF_RE.match(v)
                cell.value = (m.group(1) if is_concentrated else m.group(2)) if m else None


def _strip_macro_traces(wb) -> None:
    """.xlsm 템플릿에서 딸려오는 매크로 흔적을 제거한다.

    템플릿에는 VBA 연결(codeName)과 XLM(엑셀 4.0 매크로) defined name(`_xleta.N` 등,
    xlm="1")이 남아 있어, .xlsx 로 저장해도 Excel 이 "매크로 제외 파일이지만 매크로 사용
    내용이 들어있습니다" 경고를 띄운다. 통합문서/시트 codeName 과 xlm defined name 을 지운다.
    """
    wb.code_name = None
    for ws in wb.worksheets:
        ws.sheet_properties.codeName = None
    for name in list(wb.defined_names):
        dn = wb.defined_names[name]
        if getattr(dn, "xlm", None) or str(getattr(dn, "value", "")).startswith("#NAME?"):
            del wb.defined_names[name]


def _make_putter(ws):
    """None 은 건너뛰고 값만 기록하는 셀 setter 를 만든다(템플릿 기본 서식/라벨 보존)."""
    def put(addr, val):
        if val is not None:
            ws[addr] = val
    return put


def _fill_free(ws, data: dict) -> None:
    """Free Report 시트 채우기."""
    inputs = data.get("inputs", {}) or {}
    inter = data.get("intermediate", {}) or {}
    res = data.get("result", {}) or {}
    checks = res.get("checks", {}) or {}
    load = inputs.get("load", {}) or {}
    hull = inputs.get("hull", {}) or {}
    put = _make_putter(ws)

    # ── 좌측: 입력 / 계산 (G열) ──
    put("G15", (load.get("type") or "").capitalize())   # Load Type
    put("G16", load.get("value"))                        # Force
    put("G17", hull.get("stiffener_span_mm"))            # Length
    put("G18", inter.get("position_a_mm"))               # Position a (distributed면 None→공란)
    put("G19", inter.get("position_b_mm"))               # Position b
    put("G21", inter.get("moment_Nmm"))                  # Moment
    put("G22", inter.get("shear_N"))                     # Shear Force
    put("G24", inter.get("effective_breadth_mm"))        # Effective Breadth
    put("G25", hull.get("plate_thickness_gross_mm"))     # Thickness
    put("G26", inter.get("net_thickness_mm"))            # Net Thickness
    put("G27", inter.get("area_mm2"))                    # Area
    put("G28", inter.get("neutral_axis_mm"))             # Neutral Axis
    put("G29", inter.get("inertia_mm4"))                 # Inertia of Moment
    put("G30", inter.get("section_modulus_mm3"))         # Section Modulus

    # ── 우측: Carling Stress 패널 (L열) ──
    put("L15", inter.get("sigma_B_calc_MPa"))            # σB_Calc.
    put("L16", inter.get("sigma_B_allow_MPa"))           # σB_Allow.
    put("L17", checks.get("bending"))                    # Assessment
    put("L18", inter.get("sigma_S_calc_MPa"))            # σS_Calc.
    put("L19", inter.get("sigma_S_allow_MPa"))           # σS_Allow.
    put("L20", checks.get("shear"))                      # Assessment

    # ── 우측: Displacement 패널 ──
    put("L22", inter.get("d_calc_mm"))                   # DCalc.
    put("L23", inter.get("d_allow_mm"))                  # DAllow.(=L/500)
    put("L24", checks.get("displacement"))               # Assessment

    # ── 우측: Total ──
    put("L26", res.get("assessment"))                    # Total Assessment


def _fill_optimization(ws, data: dict) -> None:
    """Optimization Report 시트 채우기 — 최적(optimal) 카링 1개의 상세 평가."""
    inputs = data.get("inputs", {}) or {}
    inter = data.get("intermediate", {}) or {}
    load = inputs.get("load", {}) or {}
    hull = inputs.get("hull", {}) or {}
    optimal = data.get("optimal")
    if not optimal:
        raise ValueError("no feasible optimal carling — 리포트를 생성할 최적 후보가 없습니다.")

    # 최적안과 (H, T_gross) 가 일치하는 후보의 상세값을 사용
    cand = next(
        (c for c in data.get("candidates", [])
         if c.get("H_mm") == optimal.get("H_mm")
         and c.get("T_gross_mm") == optimal.get("T_gross_mm")),
        None,
    )
    if cand is None:
        raise ValueError("optimal 후보의 상세 데이터를 candidates 에서 찾지 못했습니다.")

    carling = cand.get("carling", {}) or {}
    composite = cand.get("composite", {}) or {}
    stress = cand.get("stress", {}) or {}
    disp = cand.get("displacement", {}) or {}
    checks = cand.get("checks", {}) or {}
    plate_corr = hull.get("plate_corrosion_mm") or 0
    t_gross = optimal.get("T_gross_mm")
    put = _make_putter(ws)

    # ── 좌측: 입력 (G열) ──
    put("G15", (load.get("type") or "").capitalize())    # Load Type
    put("G16", load.get("value"))                        # Force
    put("G17", hull.get("stiffener_span_mm"))            # Length
    put("G18", inter.get("position_a_mm"))               # Position a
    put("G19", inter.get("position_b_mm"))               # Position b
    put("G21", inter.get("moment_Nmm"))                  # Moment
    put("G22", inter.get("shear_N"))                     # Shear Force
    put("G24", inter.get("yield_stress_MPa"))            # Material Yield
    put("G25", optimal.get("H_mm"))                      # Depth (H)
    put("G26", t_gross)                                  # Thickness gross
    put("G27", (t_gross - plate_corr) if t_gross is not None else None)  # Net Thickness
    put("G28", carling.get("area_mm2"))                  # Area
    put("G29", carling.get("section_modulus_mm3"))       # Section Modulus
    put("G31", inputs.get("effective_breadth_mm"))       # Effective Breadth
    put("G32", hull.get("plate_thickness_gross_mm"))     # Thickness gross
    put("G33", inter.get("plate_net_thickness_mm"))      # Net Thickness
    put("G34", composite.get("area_mm2"))                # Area
    put("G35", composite.get("neutral_axis_mm"))         # Neutral Axis
    put("G36", composite.get("inertia_mm4"))             # Inertia
    put("G37", composite.get("section_modulus_mm3"))     # Section Modulus
    put("G39", cand.get("min_leg_length_mm"))            # Weld Length

    # ── 우측 패널 (L열) ──
    put("L16", carling.get("depth_per_thk"))             # Ratio Calc.(=H/T)
    put("L17", 16)                                       # Allow.
    put("L18", checks.get("depth_ratio"))                # Assessment
    put("L20", stress.get("sigma_B_calc_MPa"))           # σB_Calc.
    put("L21", stress.get("sigma_B_allow_MPa"))          # σB_Allow.
    put("L22", checks.get("bending"))                    # Assessment
    put("L23", stress.get("sigma_S_calc_MPa"))           # σS_Calc.
    put("L24", stress.get("sigma_S_allow_MPa"))          # σS_Allow.
    put("L25", checks.get("shear"))                      # Assessment
    put("L27", disp.get("d_calc_mm"))                    # DCalc.
    put("L28", disp.get("d_allow_mm"))                   # DAllow.(=L/500)
    put("L29", checks.get("displacement"))               # Assessment
    put("L31", stress.get("sigma_weld_calc_MPa"))        # σweld_Calc.
    put("L32", stress.get("sigma_weld_allow_MPa"))       # σweld_Allow.
    put("L33", checks.get("weld"))                       # Assessment
    put("L35", cand.get("assessment"))                   # Total
    put("L37", cand.get("weight_kg"))                    # Carling Weight


def generate_report(result: dict, employee_id: str) -> tuple[str, bytes]:
    """solver 전체 결과(dict)로 Carling 리포트 .xlsx bytes 를 생성한다(free/optimization 공통).

    Args:
        result: solver 출력 전체. mode 키로 free/optimization 분기.
        employee_id: 요청 사번(현재 파일명에는 미사용, 시그니처 호환 유지).

    Returns:
        (파일명, 파일 bytes).

    Raises:
        HTTPException: 템플릿 부재/로딩 실패·최적 후보 부재 등.
    """
    mode = (result.get("mode") or "free").lower()
    templates = _TEMPLATES.get(mode) or _TEMPLATES["free"]
    mode_label = _MODE_LABEL.get(mode, "Free")

    inputs = result.get("inputs") or {}
    load_type = (inputs.get("load", {}).get("type") or "concentrated").lower()
    if load_type not in templates:
        load_type = "concentrated"

    report_src = _resolve_report_dir()
    template_path = os.path.join(report_src, templates[load_type])
    if not os.path.exists(template_path):
        logger.error("Carling report template not found: %s", template_path)
        raise HTTPException(status_code=503, detail="리포트 템플릿을 찾을 수 없습니다.")

    try:
        wb = openpyxl.load_workbook(template_path, keep_links=False)
    except Exception as exc:  # DRM 암호화 등으로 평문 ZIP 이 아니면 여기서 실패
        logger.error("openpyxl failed to open template %s: %s", template_path, exc)
        raise HTTPException(
            status_code=503,
            detail="리포트 템플릿을 열 수 없습니다(템플릿이 DRM 암호화 상태일 수 있음). 서버 관리자에게 문의하세요.",
        ) from exc

    _strip_macro_traces(wb)
    ws = wb.active
    _resolve_external_formulas(ws, is_concentrated=(load_type == "concentrated"))

    try:
        if mode == "optimization":
            _fill_optimization(ws, result)
        else:
            _fill_free(ws, result)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_name = f"Carling_{mode_label}_Report_{load_type}_{timestamp}.xlsx"
    return output_name, data


# 하위 호환 alias (기존 free 엔드포인트용)
generate_free_report = generate_report
