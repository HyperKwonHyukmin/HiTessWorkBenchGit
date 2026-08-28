"""사내 DRM 대응 openpyxl.load_workbook shim.

회사 DRM 은 디스크의 .xlsx 를 at-rest 로 암호화(HHIDRMC, +4096 byte)하고 일반 로컬
프로세스는 복호화하지 못한다. 그래서 엔진이 보고서 서식 템플릿(`Report for PSA.xlsx`)을
`openpyxl.load_workbook(path)` 로 그냥 열면 BadZipFile 이 나고, 엔진은 '빈 워크북'으로
폴백해 **서식·이미지가 통째로 날아간 보고서**를 만든다.

이 shim 은 `openpyxl.load_workbook` 을 감싸서:

1. 경로 인자면 바이트를 먼저 읽고 PK(zip) 매직을 검사한다.
   - PK 이면 BytesIO 로 로드한다. 디스크 파일을 openpyxl 이 직접 열지 않으므로
     DRM 여부와 무관하게 동일하게 동작한다.
2. PK 가 아닌데 대상이 보고서 서식 템플릿이면, 번들된 `report_template.bin` 으로 폴백한다.
   .bin 은 Office 확장자가 아니라 DRM 이 건드리지 않으므로 언제나 PK 로 읽힌다.
3. 그 외에는 원래 함수를 그대로 호출해 원래 예외가 그대로 올라가게 둔다.

(엔진 FuelLine_PSA_Report.py 의 `preload_template()` / `_PRELOADED_WB` 개조를 대체한다.
 선로딩 방식은 '파이프라인 시작 시점에 이미 암호화돼 있으면' 실패했지만, 이 방식은 호출
 시점과 무관하게 번들 원본으로 폴백하므로 더 견고하다.)
"""
import io
import os
import sys

ZIP_MAGIC = b"PK\x03\x04"
BUNDLED_TEMPLATE_NAME = "report_template.bin"

# 폴백을 적용할 대상 파일명(소문자 비교). 엔진이 cwd 상대경로로 이 이름을 읽는다.
TEMPLATE_BASENAMES = ("report for psa.xlsx",)

_installed = False
_orig_load_workbook = None


def _bundled_candidates():
    """번들/배포된 report_template.bin 후보 경로를 우선순위대로 돌려준다."""
    candidates = []
    meipass = getattr(sys, "_MEIPASS", None)  # PyInstaller onefile 추출 폴더
    if meipass:
        candidates.append(os.path.join(meipass, BUNDLED_TEMPLATE_NAME))
    if getattr(sys, "frozen", False):
        candidates.append(os.path.join(os.path.dirname(sys.executable), BUNDLED_TEMPLATE_NAME))
    else:
        # 개발 실행: 어댑터 패키지 바로 위(빌드 산출 폴더)에 있을 수 있다.
        pkg_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        candidates.append(os.path.join(pkg_root, BUNDLED_TEMPLATE_NAME))
    candidates.append(os.path.join(os.getcwd(), BUNDLED_TEMPLATE_NAME))
    return candidates


def _read_if_zip(path):
    """path 를 읽어 PK 로 시작하면 bytes, 아니면 None."""
    try:
        with open(path, "rb") as f:
            data = f.read()
    except OSError:
        return None
    return data if data[:4] == ZIP_MAGIC else None


def _is_template(path):
    return os.path.basename(path).lower() in TEMPLATE_BASENAMES


def _load_workbook(filename, *args, **kwargs):
    if isinstance(filename, (str, os.PathLike)):
        path = os.fspath(filename)
        data = _read_if_zip(path)
        if data is not None:
            return _orig_load_workbook(io.BytesIO(data), *args, **kwargs)
        if _is_template(path):
            for candidate in _bundled_candidates():
                bundled = _read_if_zip(candidate)
                if bundled is not None:
                    print(f"[hitess] 서식 템플릿을 번들 사본으로 대체했습니다: {candidate}")
                    return _orig_load_workbook(io.BytesIO(bundled), *args, **kwargs)
            print("[경고][hitess] 서식 템플릿이 DRM 암호화됐거나 손상됐고 번들 사본"
                  f"({BUNDLED_TEMPLATE_NAME})도 찾지 못했습니다 — 서식 없는 보고서가 될 수 있습니다.")
    return _orig_load_workbook(filename, *args, **kwargs)


def install():
    global _installed, _orig_load_workbook
    if _installed:
        return
    import openpyxl
    import openpyxl.reader.excel as _reader

    if getattr(openpyxl.load_workbook, "_hitess_drm_shim", False):
        _installed = True
        return

    _orig_load_workbook = _reader.load_workbook
    _load_workbook._hitess_drm_shim = True
    # 엔진은 `import openpyxl` 후 `openpyxl.load_workbook(...)` 형태로 호출한다.
    # `from openpyxl.reader.excel import load_workbook` 경로도 함께 막는다.
    openpyxl.load_workbook = _load_workbook
    _reader.load_workbook = _load_workbook
    _installed = True
